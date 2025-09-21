"""State management for stateful lead generation pipeline."""

import json
import csv
from pathlib import Path
from typing import Set, Dict, Union
from openpyxl import load_workbook
from ..models.company import Company
from ..config.models import OutputConfig
from ..utils.logging import logger


class StateStore:
    """Manages state for resumable lead generation pipeline."""

    def __init__(self, output_dir: Union[str, Path], output_config: OutputConfig):
        self.output_dir = Path(output_dir)
        self.config = output_config
        self.state_dir = Path("./state")
        self.state_file = self.state_dir / "state.json"

        # In-memory state
        self.seen_companies: Set[str] = set()
        self.seen_domains: Set[str] = set()
        self.seen_emails: Set[str] = set()

        # Ensure state directory exists
        self.state_dir.mkdir(parents=True, exist_ok=True)

    def load_from_output(self) -> None:
        """Load existing state from output files and cached state."""
        try:
            # Try to load from cached state first
            if self.state_file.exists():
                with open(self.state_file, "r", encoding="utf-8") as f:
                    state_data = json.load(f)
                    self.seen_companies = set(state_data.get("companies", []))
                    self.seen_domains = set(state_data.get("domains", []))
                    self.seen_emails = set(state_data.get("emails", []))
                logger.info(
                    f"Loaded cached state: {len(self.seen_companies)} companies,"
                    f" {len(self.seen_domains)} domains, {len(self.seen_emails)} emails"
                )
                return
        except Exception as e:
            logger.warning(f"Could not load cached state: {e}")

        # If cached state not available, load from output files
        self._load_from_files()

    def _load_from_files(self) -> None:
        """Load state by reading existing output files."""
        # Load companies
        self._load_companies()
        # Load domains
        self._load_domains()
        # Load emails
        self._load_emails()

        logger.info(
            f"Loaded state from files: {len(self.seen_companies)} companies,"
            f" {len(self.seen_domains)} domains, {len(self.seen_emails)} emails"
        )

    def _load_companies(self) -> None:
        """Load existing companies from output files."""
        formats = [self.config.format.lower()]
        if self.config.format.lower() in ["json", "jsonl"]:
            formats = ["jsonl", "json"]

        for fmt in formats:
            file_path = self.output_dir / f"{self.config.companies_file}.{fmt}"
            if file_path.exists():
                try:
                    if fmt == "csv":
                        self._load_companies_csv(file_path)
                    elif fmt in ["jsonl", "json"]:
                        self._load_companies_jsonl(file_path)
                    elif fmt == "xlsx":
                        self._load_companies_xlsx(file_path)
                    else:  # txt
                        self._load_companies_txt(file_path)
                    break
                except Exception as e:
                    logger.warning(f"Error loading companies from {file_path}: {e}")

    def _normalize_company_key(self, company: Union[Company, Dict]) -> str:
        """Create normalized key for company deduplication."""
        if isinstance(company, Company):
            name = getattr(company, "name", "") or ""
            city = (
                getattr(company, "city", "") or getattr(company, "location", "") or ""
            )
            address = getattr(company, "address", "") or ""
            url = getattr(company, "url", "") or ""
        else:
            name = company.get("name", "")
            city = company.get("city", "") or company.get("location", "")
            address = company.get("address", "")
            url = company.get("url", "")

        # Normalize for deduplication (lowercase, strip whitespace)
        # Use name + location/city as primary key, with address and url as secondary
        key = f"{name.lower().strip()}|{city.lower().strip()}|{address.lower().strip()}|{url.lower().strip()}"
        return key

    def _load_companies_csv(self, file_path: Path) -> None:
        """Load companies from CSV file."""
        with open(file_path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                key = self._normalize_company_key(row)
                self.seen_companies.add(key)

    def _load_companies_jsonl(self, file_path: Path) -> None:
        """Load companies from JSONL file."""
        with open(file_path, "r", encoding="utf-8") as f:
            for line in f:
                if line.strip():
                    try:
                        company = json.loads(line.strip())
                        key = self._normalize_company_key(company)
                        self.seen_companies.add(key)
                    except json.JSONDecodeError:
                        continue

    def _load_companies_txt(self, file_path: Path) -> None:
        """Load companies from TXT file (JSON per line)."""
        self._load_companies_jsonl(file_path)

    def _load_companies_xlsx(self, file_path: Path) -> None:
        """Load companies from XLSX file."""
        try:
            wb = load_workbook(file_path)
            ws = wb.active
            if ws is None:
                return

            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                company_dict = dict(zip(headers, row))
                key = self._normalize_company_key(company_dict)
                self.seen_companies.add(key)
        except ImportError:
            logger.warning("openpyxl not available for loading XLSX companies")
        except Exception as e:
            logger.warning(f"Error loading companies from XLSX: {e}")

    def _load_domains(self) -> None:
        """Load existing domains from output files."""
        # Try TXT first, then CSV
        for ext in ["txt", "csv"]:
            file_path = self.output_dir / f"{self.config.domains_file}.{ext}"
            if file_path.exists():
                try:
                    if ext == "csv":
                        with open(file_path, "r", encoding="utf-8") as f:
                            reader = csv.reader(f)
                            next(reader, None)  # Skip header
                            for row in reader:
                                if row and row[0].strip():
                                    self.seen_domains.add(row[0].strip())
                    else:  # txt
                        with open(file_path, "r", encoding="utf-8") as f:
                            for line in f:
                                domain = line.strip()
                                if domain:
                                    self.seen_domains.add(domain)
                    break
                except Exception as e:
                    logger.warning(f"Error loading domains from {file_path}: {e}")

    def _load_emails(self) -> None:
        """Load existing emails from output files."""
        formats = [self.config.format.lower()]
        if self.config.format.lower() in ["json", "jsonl"]:
            formats = ["jsonl", "json"]

        for fmt in formats:
            file_path = self.output_dir / f"{self.config.emails_file}.{fmt}"
            if file_path.exists():
                try:
                    if fmt == "csv":
                        with open(file_path, "r", encoding="utf-8") as f:
                            reader = csv.DictReader(f)
                            for row in reader:
                                email = row.get("email", "").strip()
                                if email:
                                    self.seen_emails.add(email.lower())
                    elif fmt in ["jsonl", "json"]:
                        with open(file_path, "r", encoding="utf-8") as f:
                            for line in f:
                                if line.strip():
                                    try:
                                        email_data = json.loads(line.strip())
                                        email = email_data.get("email", "").strip()
                                        if email:
                                            self.seen_emails.add(email.lower())
                                    except json.JSONDecodeError:
                                        continue
                    elif fmt == "xlsx":
                        try:
                            wb = load_workbook(file_path)
                            ws = wb.active
                            if ws is None:
                                continue

                            headers = [cell.value for cell in ws[1]]
                            email_col = None
                            for i, header in enumerate(headers):
                                if header and "email" in header.lower():
                                    email_col = i
                                    break

                            if email_col is not None:
                                for row in ws.iter_rows(min_row=2, values_only=True):
                                    if len(row) > email_col and row[email_col]:
                                        email = str(row[email_col]).strip()
                                        if email:
                                            self.seen_emails.add(email.lower())
                        except ImportError:
                            logger.warning(
                                "openpyxl not available for loading XLSX emails"
                            )
                    break
                except Exception as e:
                    logger.warning(f"Error loading emails from {file_path}: {e}")

    def is_seen_company(self, company: Company) -> bool:
        """Check if company has been processed before."""
        key = self._normalize_company_key(company)
        return key in self.seen_companies

    def is_seen_domain(self, domain: str) -> bool:
        """Check if domain has been processed before."""
        return domain.strip() in self.seen_domains

    def is_seen_email(self, email: str) -> bool:
        """Check if email has been found before."""
        return email.lower().strip() in self.seen_emails

    def add_seen_company(self, company: Company) -> None:
        """Mark company as seen."""
        key = self._normalize_company_key(company)
        self.seen_companies.add(key)

    def add_seen_domain(self, domain: str) -> None:
        """Mark domain as seen."""
        self.seen_domains.add(domain.strip())

    def add_seen_email(self, email: str) -> None:
        """Mark email as seen."""
        self.seen_emails.add(email.lower().strip())

    def save_state(self) -> None:
        """Persist state to cache file."""
        try:
            state_data = {
                "companies": list(self.seen_companies),
                "domains": list(self.seen_domains),
                "emails": list(self.seen_emails),
            }
            with open(self.state_file, "w", encoding="utf-8") as f:
                json.dump(state_data, f, ensure_ascii=False, indent=2)
            logger.debug(f"Saved state to {self.state_file}")
        except Exception as e:
            logger.warning(f"Could not save state: {e}")

    def get_stats(self) -> Dict[str, int]:
        """Get statistics about current state."""
        return {
            "companies": len(self.seen_companies),
            "domains": len(self.seen_domains),
            "emails": len(self.seen_emails),
        }

    def clear_state(self) -> None:
        """Clear all state (for fresh runs)."""
        self.seen_companies.clear()
        self.seen_domains.clear()
        self.seen_emails.clear()
        if self.state_file.exists():
            self.state_file.unlink()
        logger.info("Cleared all state for fresh run")
