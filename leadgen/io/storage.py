import json
import csv
from pathlib import Path
from typing import List, Optional, Set
from openpyxl import Workbook
from ..models.company import Company
from ..models.email_result import EmailResult
from ..config.models import OutputConfig
from ..utils.logging import logger


class OutputManager:
    """Manages saving results (companies, domains, emails) to various formats."""

    def __init__(self, config: OutputConfig):
        self.config = config
        self.output_dir = Path(config.directory)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    # -------------------- PUBLIC --------------------

    def save_results(
        self,
        companies: List[Company],
        email_results: List[EmailResult],
        filtered_domains: Optional[Set[str]] = None,
    ):
        """Save all results to configured output format."""
        logger.info(f"Saving results to {self.output_dir}")

        # Companies
        if companies:
            self._save_companies(companies)

        # Domains
        if filtered_domains is not None:
            logger.info(filtered_domains)
            self._save_domains_filtered(filtered_domains)
        else:
            self._save_domains(companies)

        # Emails
        if email_results:
            self._save_emails(email_results)

        logger.success(f"Results saved to {self.output_dir}")

    # -------------------- COMPANIES --------------------

    def _save_companies(self, companies: List[Company]):
        """Save company data based on configured format."""
        if self.config.format.lower() in {"jsonl", "json"}:
            self._save_jsonl([c.to_dict() for c in companies], self.config.companies_file)
        elif self.config.format.lower() == "csv":
            self._save_csv([c.to_dict() for c in companies], self.config.companies_file)
        else:
            self._save_txt([c.to_dict() for c in companies], self.config.companies_file)

    # -------------------- DOMAINS --------------------

    def _save_domains(self, companies: List[Company]):
        """Extract unique valid domains from companies."""
        from ..utils.domain import DomainResolver

        resolver = DomainResolver()
        domains = {c.domain for c in companies if c.domain and resolver._is_valid_business_domain(c.domain)}
        self._save_domains_set(domains, self.config.domains_file)

    def _save_domains_filtered(self, domains: Set[str]):
        """Save pre-filtered domains from orchestrator."""
        self._save_domains_set(domains, self.config.domains_file)

    def _save_domains_set(self, domains: Set[Optional[str]], filename: str):
        """Save a set of domains in TXT or CSV."""
        if not domains:
            logger.info("No domains to save - writing empty file")
            with open(self.output_dir / f"{filename}.txt", "w", encoding="utf-8") as f:
                pass
            return

        # Remove None values
        cleaned_domains = [d for d in domains if d is not None]

        if self.config.format.lower() == "csv":
            file_path = self.output_dir / f"{filename}.csv"
            with open(file_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["domain"])
                for d in sorted(cleaned_domains):
                    writer.writerow([d])
        else:  # TXT
            file_path = self.output_dir / f"{filename}.txt"
            with open(file_path, "w", encoding="utf-8") as f:
                for d in sorted(cleaned_domains):
                    f.write(d + "\n")

        logger.info(f"Saved {len(cleaned_domains)} domains to {file_path}")

    # -------------------- EMAILS --------------------

    def _save_emails(self, email_results: List[EmailResult]):
        """Save emails in configured format."""
        emails_list = []
        for r in email_results:
            if r.success and r.emails:
                emails_cleaned = [ {k: v for k, v in e.to_dict().items() if k != "domain"} for e in r.emails ]
                emails_list.extend(emails_cleaned)

        if not emails_list:
            logger.info("No email results to save")
            return

        if self.config.format.lower() == "csv":
            self._save_csv(emails_list, self.config.emails_file)
        elif self.config.format.lower() in {"jsonl", "json"}:
            self._save_jsonl(emails_list, self.config.emails_file)
        elif self.config.format.lower() == "xlsx":
            self._save_xlsx(emails_list, self.config.emails_file)
        else:
            self._save_txt(emails_list, self.config.emails_file)

    # -------------------- FILE HELPERS --------------------

    def _save_txt(self, data: List[dict], filename: str):
        file_path = self.output_dir / f"{filename}.txt"
        with open(file_path, "a", encoding="utf-8") as f:
            for entry in data:
                f.write(json.dumps(entry, ensure_ascii=False) + "\n")
        logger.info(f"Saved {len(data)} records to {file_path}")

    def _save_csv(self, data: List[dict], filename: str):
        if not data:
            return
        file_path = self.output_dir / f"{filename}.csv"
        fieldnames = sorted({k for d in data for k in d.keys()})
        with open(file_path, "a", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(data)
        logger.info(f"Saved {len(data)} records to {file_path}")

    def _save_jsonl(self, data: List[dict], filename: str):
        file_path = self.output_dir / f"{filename}.jsonl"
        with open(file_path, "a", encoding="utf-8") as f:
            for entry in data:
                json.dump(entry, f, ensure_ascii=False)
                f.write("\n")
        logger.info(f"Saved {len(data)} records to {file_path}")

    def _save_xlsx(self, data: List[dict], filename: str):
        """Append data to XLSX file, creating a new file if it doesn't exist."""
        if not data:
            return
        file_path = self.output_dir / f"{filename}.xlsx"
        from openpyxl import load_workbook, Workbook

        if file_path.exists():
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            # Write header
            fieldnames = sorted({k for d in data for k in d.keys()})
            ws.append(fieldnames)

        fieldnames = [cell.value for cell in ws[1]]  # Use existing header
        for row in data:
            ws.append([row.get(f, "") for f in fieldnames])

        wb.save(file_path)
        logger.info(f"Appended {len(data)} records to {file_path}")
